import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
import matplotlib.dates as mdates
import os
from openpyxl import load_workbook
from openpyxl.styles import Font

# Path to the xlsx file
MOOD_TRACKING_SHEET_PATH = f'/home/jgopal/NAS/Analysis/AudioFacialEEG/Behavioral Labeling/Mood_Tracking.xlsx'
MISC_FIGURE_PATH = f'/home/jgopal/NAS/Analysis/Misc_Figures/'

# Define the date format explicitly
date_format = "%Y-%m-%d %H:%M:%S"  # Replace with your actual format

# Read the Excel file
xls = pd.ExcelFile(MOOD_TRACKING_SHEET_PATH)

# Get all sheet names that start with "S_"
sheet_names = [sheet for sheet in xls.sheet_names if sheet.startswith("S_")]

# Define the columns for analysis
columns_to_analyze = ['Mood', 'Depression', 'Anxiety']
plot_colors = {'Mood': 'blue', 'Depression': 'green', 'Anxiety': 'red'}

# Initialize a list to collect patient data for the final spreadsheet
patient_data = []

# Iterate through the sheets and analyze data
for sheet_name in sheet_names:
    df = pd.read_excel(xls, sheet_name=sheet_name)
    
    patient_info = {'Sheet Name': sheet_name}
    
    # Mood criteria for inclusion
    mood_df = df.dropna(subset=['Mood'])
    mood_df = mood_df[mood_df['Mood'].astype(bool)]

    # Convert the first column to datetime using the specified format
    mood_df[mood_df.columns[0]] = pd.to_datetime(mood_df[mood_df.columns[0]], format=date_format, errors='coerce')

    # Remove rows with invalid datetime
    mood_df = mood_df.dropna(subset=[mood_df.columns[0]])

    # Sort for chronological order
    mood_df = mood_df.sort_values(by=mood_df.columns[0])

    # Extract the datetime and mood values
    time_data_mood = mood_df[mood_df.columns[0]]
    mood_data = mood_df['Mood']

    # Calculate the metrics for mood
    num_mood_datapoints = len(mood_data)
    num_unique_mood_values = mood_data.nunique()
    mood_percent_variation = mood_data.std() / mood_data.mean() * 100 if mood_data.mean() != 0 else 0
    mood_days_span = (time_data_mood.max() - time_data_mood.min()).days

    # Add mood metrics to patient info
    patient_info['Num Mood Datapoints'] = num_mood_datapoints
    patient_info['Num Unique Mood Values'] = num_unique_mood_values
    patient_info['Mood Pct Variation'] = mood_percent_variation

    # Criteria for inclusion based on mood data
    include_patient = num_mood_datapoints >= 5 and num_unique_mood_values >= 4

    # Analyze and plot other data columns
    for column in columns_to_analyze:
        column_df = df.dropna(subset=[column])
        column_df = column_df[column_df[column].astype(bool)]

        # Convert the first column to datetime using the specified format
        column_df[column_df.columns[0]] = pd.to_datetime(column_df[column_df.columns[0]], format=date_format, errors='coerce')

        # Remove rows with invalid datetime
        column_df = column_df.dropna(subset=[column_df.columns[0]])

        # Sort for chronological order
        column_df = column_df.sort_values(by=column_df.columns[0])

        # Extract the datetime and values
        time_data = column_df[column_df.columns[0]]
        value_data = column_df[column]

        # Calculate the metrics for each column
        num_datapoints = len(value_data)
        num_unique_values = value_data.nunique()
        percent_variation = value_data.std() / value_data.mean() * 100 if value_data.mean() != 0 else 0

        # Add metrics to patient info
        patient_info[f'Num {column} Datapoints'] = num_datapoints
        patient_info[f'Num Unique {column} Values'] = num_unique_values
        patient_info[f'{column} Pct Variation'] = percent_variation

        # Plot the data
        color = plot_colors[column]
        plt.plot(time_data, value_data, marker='o', color=color, label=column)

    # Add the inclusion decision to patient info
    patient_info['Include Patient'] = 'Yes' if include_patient else 'No'

    # Append patient info to the list
    patient_data.append(patient_info)

    # Finalize the plot for this patient
    plt.title(f'{sheet_name}: Mood, Depression, and Anxiety Over Time', fontsize=20)
    plt.xlabel('Datetime', fontsize=18)
    plt.ylabel('Score (0-10)', fontsize=18)
    plt.xticks(rotation=45)
    plt.gca().xaxis.set_major_formatter(mdates.DateFormatter('%m/%d/%Y %H:%M'))
    plt.legend()
    plt.tight_layout()

    # Save the figure for this patient
    plt.savefig(os.path.join(MISC_FIGURE_PATH, f'{sheet_name}_Mood_Anxiety_Depression.png'), dpi=300)
    plt.clf()  # Clear the figure for the next patient

# Convert the patient data to a DataFrame
patient_df = pd.DataFrame(patient_data)

# Save the DataFrame to an Excel file
new_file_path = os.path.join(os.path.dirname(MOOD_TRACKING_SHEET_PATH), 'Mood_Tracking_Overview.xlsx')
with pd.ExcelWriter(new_file_path, engine='openpyxl') as writer:
    patient_df.to_excel(writer, sheet_name='Data_Overview', index=False)

# Load the workbook and access the sheet
wb = load_workbook(new_file_path)
ws = wb['Data_Overview']

# Bold the rows where 'Include Patient' is 'Yes'
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    if row[0].value == 'Yes':
        for cell in row:
            cell.font = Font(bold=True)

# Save the updated workbook
wb.save(new_file_path)
