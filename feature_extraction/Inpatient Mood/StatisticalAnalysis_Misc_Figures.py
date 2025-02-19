import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
import matplotlib.dates as mdates
import os
from openpyxl import load_workbook
from openpyxl.styles import Font
import itertools
import re

# Path to the xlsx file
MOOD_TRACKING_SHEET_PATH = '/home/jgopal/NAS/Analysis/AudioFacialEEG/Behavioral Labeling/Mood_Tracking.xlsx'
MISC_FIGURE_PATH = '/home/jgopal/NAS/Analysis/Misc_Figures/'

# 1) Read the Excel, find all sheets that start with "S_"
xls = pd.ExcelFile(MOOD_TRACKING_SHEET_PATH)
sheet_names = [sheet for sheet in xls.sheet_names if sheet.startswith("S_")]

###############################################################################
# 2) Dynamically discover all metrics
#    Skip: first column, "notes"/"datetime", columns containing "catch trial",
#          columns with "p" followed by digits anywhere in name.
###############################################################################
all_metrics = set()
for sheet_name in sheet_names:
    df = pd.read_excel(xls, sheet_name=sheet_name)
    if df.empty:
        continue
    
    # Exclude the first column (assumed datetime)
    candidate_cols = df.columns[1:]
    
    for col in candidate_cols:
        col_lower = col.lower().strip()
        
        # Conditions for ignoring this column:
        #   - If it's named "notes" or "datetime"
        #   - If it contains "catch trial"
        #   - If it matches pattern p\d+ anywhere in the name
        if col_lower == 'notes':
            continue
        if col_lower == 'datetime':
            continue
        if 'catch trial' in col_lower:
            continue
        if 'panas' in col_lower:
            continue
        if 'unnamed' in col_lower:
            continue
        if re.search(r'p\d+', col_lower):
            continue
        
        all_metrics.add(col)

# Convert set to a sorted list
all_metrics = sorted(all_metrics)

###############################################################################
# 3) Prepare color cycling so each metric has a distinct color in the same plot
###############################################################################
color_cycle = itertools.cycle(plt.cm.tab10(np.linspace(0, 1, 10)))  
metric_color_map = {}
for metric in all_metrics:
    metric_color_map[metric] = next(color_cycle)

###############################################################################
# 4) Inclusion criteria and plotting per sheet
###############################################################################
patient_data = []

for sheet_name in sheet_names:
    df = pd.read_excel(xls, sheet_name=sheet_name)
    patient_info = {'Sheet Name': sheet_name}
    
    # Convert the first column to datetime
    if not df.empty:
        if not pd.api.types.is_datetime64_any_dtype(df[df.columns[0]]):
            df[df.columns[0]] = pd.to_datetime(df[df.columns[0]], errors='coerce')
        df = df.dropna(subset=[df.columns[0]])
        df = df.sort_values(by=df.columns[0])

    # Start a plot for this patient (overlay all metrics)
    plt.figure(figsize=(10, 6))
    
    for metric in all_metrics:
        # If metric doesn't exist, set default stats and skip
        if metric not in df.columns:
            patient_info[f'Num Datapoints - {metric}'] = 0
            patient_info[f'Num Unique - {metric}'] = 0
            patient_info[f'Range - {metric}'] = 0
            patient_info[f'Is Included - {metric}'] = 'No'
            continue
        
        # Filter out rows with NaN or zero (False) in this metric
        metric_df = df.dropna(subset=[metric])
        metric_df = metric_df[metric_df[metric].astype(bool)]
        if metric_df.empty:
            patient_info[f'Num Datapoints - {metric}'] = 0
            patient_info[f'Num Unique - {metric}'] = 0
            patient_info[f'Range - {metric}'] = 0
            patient_info[f'Is Included - {metric}'] = 'No'
            continue
        
        # Ensure chronological order
        metric_df = metric_df.sort_values(by=df.columns[0])
        
        # Calculate stats
        values = metric_df[metric]
        time_data = metric_df[df.columns[0]]
        
        num_datapoints = len(values)
        num_unique = values.nunique()
        val_range = values.max() - values.min()
        
        patient_info[f'Num Datapoints - {metric}'] = num_datapoints
        patient_info[f'Num Unique - {metric}'] = num_unique
        patient_info[f'Range - {metric}'] = val_range
        
        # Check inclusion (>=5 datapoints, >=3 unique, range>=5)
        if (num_datapoints >= 5) and (num_unique >= 3) and (val_range >= 5):
            patient_info[f'Is Included - {metric}'] = 'Yes'
        else:
            patient_info[f'Is Included - {metric}'] = 'No'
        
        # Plot
        plt.plot(time_data, values,
                 marker='o',
                 color=metric_color_map[metric],
                 label=metric)
    
    plt.title(f'{sheet_name}: All Metrics Over Time', fontsize=16)
    plt.xlabel('Datetime', fontsize=14)
    plt.ylabel('Score', fontsize=14)
    plt.gca().xaxis.set_major_formatter(mdates.DateFormatter('%m/%d/%Y %H:%M'))
    plt.xticks(rotation=45)
    plt.legend()
    plt.tight_layout()
    
    # Save figure
    out_path = os.path.join(MISC_FIGURE_PATH, f'{sheet_name}_Metrics_Over_Time.png')
    plt.savefig(out_path, dpi=300)
    plt.clf()
    
    patient_data.append(patient_info)

###############################################################################
# 5) Save summary stats to Excel
###############################################################################
patient_df = pd.DataFrame(patient_data)
new_file_path = os.path.join(
    os.path.dirname(MOOD_TRACKING_SHEET_PATH),
    'Mood_Tracking_Overview.xlsx'
)

with pd.ExcelWriter(new_file_path, engine='openpyxl') as writer:
    patient_df.to_excel(writer, sheet_name='Data_Overview', index=False)

# Optionally bold entire rows if any metric is 'Yes'
wb = load_workbook(new_file_path)
ws = wb['Data_Overview']

include_cols = [c for c in patient_df.columns if c.startswith('Is Included - ')]
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    # Check if any "Is Included - {metric}" cell is 'Yes'
    if any(cell.value == 'Yes' for cell in row if ws.cell(row=1, column=cell.column).value in include_cols):
        for cell in row:
            cell.font = Font(bold=True)

wb.save(new_file_path)




###############################################################################
# 0) Paths and constants
###############################################################################

# Path where your overview Excel is stored. You can adapt this if needed.
OVERVIEW_EXCEL_PATH = new_file_path

# Create a subfolder named "Summary_Stats" in the same directory as the overview Excel
BASE_FOLDER = os.path.dirname(OVERVIEW_EXCEL_PATH)
SUMMARY_FIGURES_PATH = os.path.join(BASE_FOLDER, "Summary_Stats")

# Make sure the subfolder exists (creates if missing)
os.makedirs(SUMMARY_FIGURES_PATH, exist_ok=True)

###############################################################################
# 1) Read the overview data
###############################################################################

df_overview = pd.read_excel(OVERVIEW_EXCEL_PATH)

# Example columns in df_overview for each metric "Mood":
# - "Num Datapoints - Mood"
# - "Num Unique - Mood"
# - "Range - Mood"
# - "Is Included - Mood"
#
# We also have "Sheet Name" for each patient.

###############################################################################
# 2) Extract the list of metrics from the columns
#    (We look for columns that match "Num Datapoints - XYZ", etc.)
###############################################################################

pattern_metric = re.compile(r"^Num Datapoints - (.+)$")  
# This pattern captures the part after "Num Datapoints - " as the metric name.

metrics = []
for col in df_overview.columns:
    match = pattern_metric.match(col)
    if match:
        metrics.append(match.group(1))

metrics = sorted(set(metrics))  # unique, sorted metric names

###############################################################################
# 3) Create a helper function to parse "S_###" or "S_###something"
#    and return the numeric part (e.g., "199" -> 199).
###############################################################################

def extract_numeric_suffix(sheet_name):
    """
    Extracts the integer that appears after 'S_' at the start of sheet_name.
    For example:
      "S_199" -> 199
      "S_202b" -> 202
    If no recognizable integer is found, returns None.
    """
    # If it strictly starts with S_ we take the substring after that
    # Then we parse out any leading digits
    match = re.match(r"^S_(\d+)", sheet_name)
    if match:
        return int(match.group(1))
    else:
        return None

###############################################################################
# 4) Summaries for each metric
###############################################################################

summary_records = []
for metric in metrics:
    # Build column names
    col_datapoints = f"Num Datapoints - {metric}"
    col_range = f"Range - {metric}"
    col_included = f"Is Included - {metric}"
    
    # Subset the DF for patients that actually have >0 data for this metric
    # (meaning "Num Datapoints - metric" > 0)
    mask_nonzero = df_overview[col_datapoints] > 0
    df_nonzero = df_overview[mask_nonzero].copy()
    
    # Among these, how many are "included"?
    # Note: a patient is included if "Is Included - metric" == "Yes"
    mask_included = (df_nonzero[col_included] == "Yes")
    df_included = df_nonzero[mask_included].copy()
    
    # 1) Number of patients included
    num_included = len(df_included)
    
    # 2) Among included patients: mean/median # of datapoints, mean/median range
    if num_included > 0:
        mean_datapoints_included = df_included[col_datapoints].mean()
        median_datapoints_included = df_included[col_datapoints].median()
        mean_range_included = df_included[col_range].mean()
        median_range_included = df_included[col_range].median()
    else:
        mean_datapoints_included = 0
        median_datapoints_included = 0
        mean_range_included = 0
        median_range_included = 0
    
    # 3) Total number of patients with non-zero scores
    num_nonzero = len(df_nonzero)
    
    summary_records.append({
        "Metric": metric,
        "Num Patients Included": num_included,
        "Mean Datapoints (Included)": mean_datapoints_included,
        "Median Datapoints (Included)": median_datapoints_included,
        "Mean Range (Included)": mean_range_included,
        "Median Range (Included)": median_range_included,
        "Total Patients w/ Non-Zero": num_nonzero
    })

# Convert to DataFrame for easy viewing
df_summary = pd.DataFrame(summary_records)

###############################################################################
# 5) Print or save summary to CSV (optional)
###############################################################################

print("===== Summary Statistics by Metric =====")
print(df_summary)

summary_csv_path = os.path.join(SUMMARY_FIGURES_PATH, "Metric_Summary_Stats.csv")
df_summary.to_csv(summary_csv_path, index=False)

###############################################################################
# 6) Make a "nice concise bar or box plot" for each metric
#    showing distribution of (# datapoints) or (range) for:
#      (a) all patients with non-zero data
#      (b) included patients only
###############################################################################

# We'll make 2 subplots:
#   Left: distribution of # datapoints
#   Right: distribution of range
# We do a box plot with two boxes each: "All Non-Zero" vs. "Included".
#
# Then we save each figure as: "{metric}_Distribution.png"

for metric in metrics:
    col_datapoints = f"Num Datapoints - {metric}"
    col_range = f"Range - {metric}"
    col_included = f"Is Included - {metric}"
    
    # All non-zero
    mask_nonzero = df_overview[col_datapoints] > 0
    data_all_datapoints = df_overview.loc[mask_nonzero, col_datapoints]
    data_all_range = df_overview.loc[mask_nonzero, col_range]
    
    # Included subset
    mask_included = (mask_nonzero & (df_overview[col_included] == "Yes"))
    data_incl_datapoints = df_overview.loc[mask_included, col_datapoints]
    data_incl_range = df_overview.loc[mask_included, col_range]

    # Create figure
    fig, axes = plt.subplots(1, 2, figsize=(10, 5))
    fig.suptitle(f"Distribution for Metric: {metric}", fontsize=16)
    
    # Left subplot: # of datapoints
    axes[0].boxplot(
        [data_all_datapoints.dropna(), data_incl_datapoints.dropna()],
        labels=["All Non-Zero", "Included"]
    )
    axes[0].set_title("Num Datapoints")
    
    # Right subplot: range
    axes[1].boxplot(
        [data_all_range.dropna(), data_incl_range.dropna()],
        labels=["All Non-Zero", "Included"]
    )
    axes[1].set_title("Range")
    
    plt.tight_layout()
    
    # Save figure
    outname = f"{metric}_Distribution.png"
    outpath = os.path.join(SUMMARY_FIGURES_PATH, outname)
    plt.savefig(outpath, dpi=300)
    plt.close(fig)

###############################################################################
# 7) For each metric, make a plot of # datapoints vs. the numeric suffix of sheet
#    We'll call that suffix "Patient Recency."
#    Save as "Metric_Recency_Plot_{metric}.png".
###############################################################################

# First, let's parse the numeric suffix for each patient (Sheet Name).
df_overview["Sheet Numeric"] = df_overview["Sheet Name"].apply(extract_numeric_suffix)

# We’ll store the numeric suffix in a new column "Sheet Numeric" to help with sorting.

for metric in metrics:
    col_datapoints = f"Num Datapoints - {metric}"
    
    # Consider only patients who have at least one datapoint (non-zero).
    sub = df_overview[df_overview[col_datapoints] > 0].copy()
    
    # If "Sheet Numeric" is None for some reason, we exclude them from the plot
    sub = sub.dropna(subset=["Sheet Numeric"])
    
    # Sort by numeric suffix ascending
    sub = sub.sort_values("Sheet Numeric")
    
    # X = sorted numeric suffix
    x_vals = np.arange(len(sub))  # 0,1,2,... for however many we have
    y_vals = sub[col_datapoints].values
    
    # For labeling or debugging, we might store their actual suffix in text
    # but we won't put it on the axis if you don't want to crowd it.
    
    # Plot
    plt.figure(figsize=(8, 5))
    plt.plot(x_vals, y_vals, marker='o', linestyle='--')
    plt.axhline(y=5, color='red', linestyle='--', linewidth=1.5)
    plt.title(f"# of Datapoints over Patient Recency for {metric}")
    plt.xlabel("Patient Recency (sorted by sheet numeric)")
    plt.ylabel("Num Self-Reports")
    
    # Optional: If you'd like some x-tick labels showing the actual S_###:
    #   labels = sub["Sheet Name"].values
    #   plt.xticks(x_vals, labels, rotation=45)
    # Otherwise, we leave them blank or minimal.
    
    plt.tight_layout()
    
    outname = f"Metric_Recency_Plot_{metric}.png"
    outpath = os.path.join(SUMMARY_FIGURES_PATH, outname)
    plt.savefig(outpath, dpi=300)
    plt.close()